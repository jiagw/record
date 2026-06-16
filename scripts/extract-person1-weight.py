"""从 Excel 饮食与运动记录表提取人员1体重数据，输出 JSON。"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path(
    r"C:\Users\jgq\Desktop\2026年5月9日 - 2026年6月8日 饮食与运动记录表（共31天）.xlsx"
)
OUTPUT = Path(__file__).resolve().parents[1] / "src" / "data" / "person1-weight.json"
DATE_COL = 1
PERSON1_WEIGHT_COL = 6


def extract_person1_weights(xlsx_path: Path) -> list[dict[str, float | str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows: list[dict[str, float | str]] = []

    for row_idx in range(2, ws.max_row + 1):
        date_value = ws.cell(row_idx, DATE_COL).value
        weight_value = ws.cell(row_idx, PERSON1_WEIGHT_COL).value

        if date_value is None:
            continue
        if not isinstance(weight_value, (int, float)):
            continue

        if isinstance(date_value, datetime):
            date_str = date_value.strftime("%Y-%m-%d")
        else:
            date_str = str(date_value)

        rows.append({"date": date_str, "weight": round(float(weight_value), 1)})

    return rows


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT

    rows = extract_person1_weights(xlsx_path)
    output_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"extracted {len(rows)} records -> {output_path}")


if __name__ == "__main__":
    main()
